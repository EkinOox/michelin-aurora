#!/usr/bin/env python3
"""
seed_from_xlsx.py — Importe les pneus et revendeurs depuis les fichiers XLSX
dans la base de données PostgreSQL via Docker.

Usage:
    python3 scripts/seed_from_xlsx.py [--tires] [--retailers] [--dry-run]

Sans argument, importe les deux. --dry-run affiche le SQL sans l'exécuter.

Prérequis :
    pip3 install openpyxl   (ou : /usr/bin/python3 si déjà installé)
"""

from __future__ import annotations

import argparse
import datetime
import hashlib
import re
import subprocess
import sys
import uuid
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    sys.exit("❌  openpyxl manquant. Installez-le : pip3 install openpyxl")

# ── Chemins ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent
CATALOG_XLSX  = ROOT / "2W Bicycle Product Catalog v4 - 2026.xlsx"
RETAILER_XLSX = ROOT / "Liste Retail MICHELIN.xlsx"

# ── Namespace UUID déterministe (reproductible d'un run à l'autre) ────────────
NS = uuid.UUID("6ba7b812-9dad-11d1-80b4-00c04fd430c8")

# ── Correspondances CYCLE TYPE WEB → bike_type Symfony ───────────────────────
BIKE_TYPE_MAP = {
    "ROAD":                       "route",
    "ROAD,E-BIKE":                "route",
    "MTB":                        "vtt",
    "MTB,E-BIKE":                 "vtt",
    "GRAVEL,COMMUTING & TOUR,E-BIKE": "gravel",
    "GRAVEL,E-BIKE":              "gravel",
    "COMMUTING & TOUR":           "vae",
    "COMMUTING & TOUR,E-BIKE":    "vae",
    "E-BIKE":                     "vae",
}

# Types ignorés (tubes, enfants)
SKIP_TYPES = {"INNER TUBES", "KIDS"}

COLOR_TOKEN = {
    "route":  "#27509B",
    "vtt":    "#E71D36",
    "gravel": "#84BD00",
    "vae":    "#2EC4B6",
}

SEGMENT_PRICE = {
    "PREMIUM RACING LINE":      (95,  140),
    "PREMIUM COMPETITION LINE": (70,  100),
    "PREMIUM PERFORMANCE LINE": (45,   80),
    "ACCESS LINE":              (20,   50),
}

SEGMENT_KM = {
    "PREMIUM RACING LINE":      4000,
    "PREMIUM COMPETITION LINE": 5000,
    "PREMIUM PERFORMANCE LINE": 6000,
    "ACCESS LINE":              8000,
}

SEGMENT_TAG = {
    "PREMIUM RACING LINE":      "Racing",
    "PREMIUM COMPETITION LINE": "Competition",
    "PREMIUM PERFORMANCE LINE": "Performance",
    "ACCESS LINE":              "Access",
}

# ─────────────────────────────────────────────────────────────────────────────

def det_price(segment: str, name: str) -> float:
    """Prix déterministe dans la fourchette du segment."""
    lo, hi = SEGMENT_PRICE.get(segment, (30, 60))
    h = int(hashlib.md5(name.encode()).hexdigest()[:4], 16)
    return round(lo + (h / 0xFFFF) * (hi - lo), 2)


def _num(val) -> Optional[float]:
    """Convertit une valeur en float, None si impossible."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def diameter_label(cycle_web: str, diam_mm, diam_in, width_mm, width_in) -> Optional[str]:
    """Formate le label de diamètre/largeur."""
    dm, di = _num(diam_mm), _num(diam_in)
    wm, wi = _num(width_mm), _num(width_in)
    if "MTB" in cycle_web or "GRAVEL" in cycle_web:
        d = di or (round(dm / 25.4) if dm else None)
        w = wi or (round(wm / 25.4, 1) if wm else None)
        if d and w:
            return f'{int(d)}" × {w}"'
    else:
        d = dm or (round(di * 25.4) if di else None)
        w = wm or (round(wi * 25.4) if wi else None)
        if d and w:
            return f"{int(d)}C × {int(w)}mm"
    return None


def clean_name(raw: str) -> str:
    """Retire le suffixe de ligne marketing du nom de la gamme."""
    for suffix in [
        " RACING LINE", " COMPETITION LINE", " PERFORMANCE LINE", " ACCESS LINE",
    ]:
        raw = raw.replace(suffix, "")
    # Retire le préfixe "MICHELIN "
    raw = re.sub(r"^MICHELIN\s+", "", raw, flags=re.IGNORECASE)
    return raw.strip().title()


def clean_subtitle(use: str | None, terrain: str | None) -> str | None:
    """Construit un sous-titre lisible."""
    parts = []
    if use:
        parts.append(use.split(",")[0].title())
    if terrain:
        parts.append(terrain.split(",")[0].title())
    return " · ".join(parts) if parts else None


def esc(val: str) -> str:
    """Échappe les apostrophes pour SQL."""
    return str(val).replace("'", "''")


# ── Parsers ───────────────────────────────────────────────────────────────────

def parse_tires() -> list[dict]:
    wb = openpyxl.load_workbook(CATALOG_XLSX)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers) if h}

    seen_ranges: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        range_name = row[idx.get("Web Range Name", -1)]
        if not range_name:
            continue

        # Skip if discontinued date is set and already in the past
        disc = row[idx.get("Discontinued Date")]
        if disc and isinstance(disc, datetime.datetime) and disc < datetime.datetime.now():
            continue

        cycle_web = str(row[idx.get("CYCLE TYPE WEB", -1)] or "").strip()

        if cycle_web in SKIP_TYPES or not cycle_web:
            continue

        bike_type = BIKE_TYPE_MAP.get(cycle_web)
        if not bike_type:
            continue

        # One representative entry per range (first occurrence)
        if range_name in seen_ranges:
            continue

        segment  = str(row[idx.get("Segment", -1)] or "PREMIUM PERFORMANCE LINE")
        weight   = row[idx.get("Weight (g)", -1)]
        diam_mm  = row[idx.get("Web Diameter (mm)", -1)]
        diam_in  = row[idx.get("Web Diameter (Inch)", -1)]
        width_mm = row[idx.get("Web Width (mm)", -1)]
        width_in = row[idx.get("Web Width (Inch)", -1)]
        use      = row[idx.get("Use", -1)]
        terrain  = row[idx.get("Terrain Types", -1)]

        clean = clean_name(range_name)
        dlabel = diameter_label(cycle_web, diam_mm, diam_in, width_mm, width_in)
        subtitle = clean_subtitle(
            str(use) if use else None,
            str(terrain) if terrain else None,
        )

        seen_ranges[range_name] = {
            "id":             str(uuid.uuid5(NS, range_name)),
            "name":           clean,
            "bike_type":      bike_type,
            "price_eur":      det_price(segment, range_name),
            "avg_km_lifetime": SEGMENT_KM.get(segment, 5000),
            "subtitle":       subtitle,
            "tag":            SEGMENT_TAG.get(segment, "Performance"),
            "color_token":    COLOR_TOKEN[bike_type],
            "weight_g":       int(weight) if weight and not isinstance(weight, str) else None,
            "diameter_label": dlabel,
        }

    return list(seen_ranges.values())


def parse_retailers() -> list[dict]:
    COUNTRY_NAMES = {
        "UK": ("🇬🇧", "Royaume-Uni"),
        "DE": ("🇩🇪", "Allemagne"),
        "ES": ("🇪🇸", "Espagne"),
        "NL": ("🇳🇱", "Pays-Bas"),
        "IT": ("🇮🇹", "Italie"),
        "PL": ("🇵🇱", "Pologne"),
        "BE": ("🇧🇪", "Belgique"),
        "FR": ("🇫🇷", "France"),
    }

    wb = openpyxl.load_workbook(RETAILER_XLSX)
    ws = wb.active

    retailers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        _region, country, url = row[0], row[1], row[2]
        if not url or not country:
            continue
        url = str(url).strip()
        if not url.startswith("http"):
            url = "https://" + url
        domain = re.sub(r"https?://(www\.)?", "", url).rstrip("/")
        name = domain.split(".")[0].capitalize()
        flag, country_name = COUNTRY_NAMES.get(str(country).strip(), ("🌍", str(country)))
        retailers.append({
            "id":   str(uuid.uuid5(NS, url)),
            "name": name,
            "sub":  f"{flag} {country_name}",
            "url":  url,
        })

    return retailers


# ── SQL builders ──────────────────────────────────────────────────────────────

def tires_to_sql(tires: list[dict]) -> str:
    lines = []
    for t in tires:
        w   = t["weight_g"] if t["weight_g"] is not None else "NULL"
        dl  = f"'{esc(t['diameter_label'])}'" if t["diameter_label"] else "NULL"
        sub = f"'{esc(t['subtitle'])}'"        if t["subtitle"]       else "NULL"
        lines.append(
            f"INSERT INTO tires "
            f"(id, name, bike_type, price_eur, avg_km_lifetime, subtitle, tag, color_token, weight_g, diameter_label) "
            f"VALUES ("
            f"'{t['id']}', '{esc(t['name'])}', '{t['bike_type']}', "
            f"{t['price_eur']}, {t['avg_km_lifetime']}, "
            f"{sub}, '{esc(t['tag'])}', '{esc(t['color_token'])}', {w}, {dl}"
            f") ON CONFLICT (id) DO UPDATE SET "
            f"name=EXCLUDED.name, bike_type=EXCLUDED.bike_type, "
            f"price_eur=EXCLUDED.price_eur, avg_km_lifetime=EXCLUDED.avg_km_lifetime, "
            f"subtitle=EXCLUDED.subtitle, tag=EXCLUDED.tag, "
            f"color_token=EXCLUDED.color_token, weight_g=EXCLUDED.weight_g, "
            f"diameter_label=EXCLUDED.diameter_label;"
        )
    return "\n".join(lines)


def retailers_to_sql(retailers: list[dict]) -> str:
    lines = []
    for r in retailers:
        lines.append(
            f"INSERT INTO retailers (id, name, sub, url) "
            f"VALUES ('{r['id']}', '{esc(r['name'])}', '{esc(r['sub'])}', '{esc(r['url'])}') "
            f"ON CONFLICT (id) DO UPDATE SET "
            f"name=EXCLUDED.name, sub=EXCLUDED.sub, url=EXCLUDED.url;"
        )
    return "\n".join(lines)


# ── Exécution ─────────────────────────────────────────────────────────────────

def run_sql(sql: str, dry_run: bool) -> None:
    if dry_run:
        print(sql)
        return

    compose_file = ROOT / "docker-compose.yml"
    proc = subprocess.run(
        ["docker", "compose", "-f", str(compose_file),
         "exec", "-T", "db",
         "psql", "-U", "aurora", "-d", "aurora"],
        input=sql.encode(),
        capture_output=True,
    )
    if proc.returncode != 0:
        print("❌  Erreur psql :", proc.stderr.decode(), file=sys.stderr)
        sys.exit(1)

    # Compte les lignes INSERT/UPDATE dans la sortie
    out = proc.stdout.decode()
    inserted = out.count("INSERT 0 1")
    updated  = out.count("UPDATE 1")
    print(f"   ✓ {inserted} inserted, {updated} updated")


def main():
    parser = argparse.ArgumentParser(description="Seed Aurora DB from XLSX files")
    parser.add_argument("--tires",     action="store_true", help="Importer uniquement les pneus")
    parser.add_argument("--retailers", action="store_true", help="Importer uniquement les revendeurs")
    parser.add_argument("--dry-run",   action="store_true", help="Afficher le SQL sans l'exécuter")
    args = parser.parse_args()

    # Par défaut : les deux
    do_tires     = args.tires     or not (args.tires or args.retailers)
    do_retailers = args.retailers or not (args.tires or args.retailers)

    if do_tires:
        if not CATALOG_XLSX.exists():
            print(f"⚠️   Fichier introuvable : {CATALOG_XLSX}")
        else:
            print(f"📦  Lecture du catalogue pneus…")
            tires = parse_tires()
            print(f"   {len(tires)} gammes actives trouvées")
            sql = tires_to_sql(tires)
            print(f"{'[DRY-RUN] ' if args.dry_run else ''}Insertion des pneus…")
            run_sql(sql, args.dry_run)

    if do_retailers:
        if not RETAILER_XLSX.exists():
            print(f"⚠️   Fichier introuvable : {RETAILER_XLSX}")
        else:
            print(f"🏪  Lecture de la liste revendeurs…")
            retailers = parse_retailers()
            print(f"   {len(retailers)} revendeurs trouvés")
            sql = retailers_to_sql(retailers)
            print(f"{'[DRY-RUN] ' if args.dry_run else ''}Insertion des revendeurs…")
            run_sql(sql, args.dry_run)

    if not args.dry_run:
        print("\n✅  Import terminé.")


if __name__ == "__main__":
    main()
